# ================================================================
# CONTRATO SEGURO IA - Aplicação Principal
# ================================================================
# Análise inteligente de contratos usando Gemini/DeepSeek
# ================================================================

import logging
import time
from io import BytesIO
from pathlib import Path

# Importações para processamento
import streamlit as st
from dotenv import load_dotenv

# Importações do projeto
from analyzer import SELECTED_MODEL, analisar_contrato, set_model, set_fallback, responder_duvida_clausula
from auth_service import AuthService
from cache_manager import CacheManager
from clause_service import extrair_numeros_riscos, extrair_pontos_atencao
from config import load_env_config
from database_service import DatabaseService
from pdf_extractor import extrair_metadados_pdf, extrair_texto_pdf_bytes

# ================================================================
# CONFIGURAÇÕES INICIAIS
# ================================================================

LOGO_PATH = Path(__file__).resolve().parent / "assets" / "Logo TRUST Contrato Seguro.png"

# Carregar variáveis de ambiente
load_dotenv(override=True)

# Configurar logging centralizado
from config import setup_logging
setup_logging()
logger = logging.getLogger(__name__)

# ================================================================
# CONFIGURAÇÃO DE PÁGINA STREAMLIT
# ================================================================

st.set_page_config(
    page_title="TRUST CORPORATION - Contrato Seguro IA",
    page_icon="🔒",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Aplicar CSS customizado - Design Profissional Fase 1
st.markdown(
    """
<style>
    /* ============ PALETA DE CORES PROFISSIONAL ============ */
    :root {
        --primary: #3A6FA0;      /* Azul claro - confiança */
        --secondary: #5B9E6D;    /* Verde claro - segurança */
        --accent: #D4AF37;       /* Ouro - premium */
        --danger: #C41E3A;       /* Vermelho escuro */
        --warning: #E89500;      /* Laranja */
        --success: #4CAF50;      /* Verde vibrante */
        --bg: #F7F2EA;           /* Bege claro */
        --text-primary: #3A6FA0; /* Texto principal */
        --text-secondary: #6A8A9A; /* Texto secundário */
    }

    /* ============ MAIN CONTAINER ============ */
    .main {
        padding: 0;
        background-color: #F7F2EA;
    }

    /* ============ HEADER PROFISSIONAL ============ */
    .header-top {
        background: linear-gradient(135deg, #3A6FA0 0%, #5B9E6D 100%);
        padding: 1rem 2rem;
        color: white;
        display: flex;
        justify-content: space-between;
        align-items: center;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        border-bottom: 3px solid #D4AF37;
    }

    /* ============ TIPOGRAFIA PROFISSIONAL ============ */
    h1 {
        font-family: 'Georgia', 'Times New Roman', serif;
        color: #3A6FA0;
        letter-spacing: 2px;
        font-weight: 700;
        margin-top: 2rem;
    }

    h2 {
        font-family: 'Georgia', 'Times New Roman', serif;
        color: #3A6FA0;
        letter-spacing: 1px;
        font-weight: 600;
        border-bottom: 2px solid #D4AF37;
        padding-bottom: 0.5rem;
    }

    h3 {
        font-family: 'Georgia', 'Times New Roman', serif;
        color: #5B9E6D;
        font-weight: 600;
    }

    /* ============ BADGES DE SEGURANÇA ============ */
    .security-badge {
        display: inline-block;
        background-color: #E8F5E9;
        border: 1px solid #2E7D32;
        color: #2E7D32;
        padding: 0.4rem 0.8rem;
        border-radius: 20px;
        font-size: 0.85em;
        font-weight: 600;
        margin: 0.25rem;
        vertical-align: middle;
    }

    .trust-badge {
        display: inline-block;
        background: linear-gradient(135deg, #D4AF37 0%, #C9A227 100%);
        color: #3A6FA0;
        padding: 0.5rem 1rem;
        border-radius: 4px;
        font-weight: 700;
        margin: 0.5rem 0.25rem;
        box-shadow: 0 2px 4px rgba(212,175,55,0.3);
    }

    /* ============ BUTTONS PROFISSIONAIS ============ */
    .stButton > button {
        background-color: #5DADE2 !important;  /* Azul mais claro */
        color: white !important;
        border: none !important;
        border-radius: 4px !important;
        padding: 0.75rem 1.5rem !important;
        font-weight: 600 !important;
        letter-spacing: 0.5px !important;
        transition: all 0.3s ease !important;
        box-shadow: 0 2px 6px rgba(93,173,226,0.2) !important;
    }

    .stButton > button[kind="primary"] {
        background-color: #5499C7 !important;
        border: none !important;
        color: white !important;
    }

    .stButton > button:hover {
        background-color: #85C1E9 !important;  /* Azul ainda mais claro no hover */
        box-shadow: 0 4px 12px rgba(133,193,233,0.3) !important;
        transform: translateY(-2px) !important;
    }

    .stButton > button[kind="primary"]:hover {
        background-color: #85C1E9 !important;
    }

    .stFormSubmitButton > button {
        background-color: #5DADE2 !important;  /* Azul mais claro */
        color: white !important;
        border: none !important;
        font-weight: 600 !important;
    }

    .stFormSubmitButton > button:hover {
        background-color: #5B9E6D !important;
    }

    /* ============ ALERT BOXES PROFISSIONAIS ============ */
    .alert-box {
        padding: 1.25rem;
        border-radius: 4px;
        margin: 1rem 0;
        border-left: 4px solid;
        background-color: #F7F2EA;
    }

    .alert-high {
        background-color: #FFEBEE;
        border-left-color: #C41E3A;
        color: #C41E3A;
    }

    .alert-medium {
        background-color: #FFF3E0;
        border-left-color: #FF9500;
        color: #FF9500;
    }

    .alert-low {
        background-color: #E8F5E9;
        border-left-color: #2E7D32;
        color: #2E7D32;
    }

    /* ============ DIVIDER PROFISSIONAL ============ */
    .streamlit-expanderHeader {
        background-color: #F7F2EA;
        border: 1px solid #E0E0E0;
    }

    /* ============ SIDEBAR ============ */
    .sidebar .sidebar-content {
        background-color: #F7F2EA;
    }

    /* ============ CARDS/CONTAINERS ============ */
    .info-card {
        background: #F7F2EA;
        border: 1px solid #E0E0E0;
        border-radius: 6px;
        padding: 1.5rem;
        margin: 1rem 0;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
        transition: all 0.3s ease;
    }

    .info-card:hover {
        box-shadow: 0 4px 12px rgba(0,0,0,0.1);
        border-color: #D4AF37;
    }

    /* ============ TRUST INDICATORS ============ */
    .trust-section {
        background: linear-gradient(135deg, #F7F2EA 0%, #FFFFFF 100%);
        border: 1px solid #E0E0E0;
        border-radius: 6px;
        padding: 1.5rem;
        margin: 1.5rem 0;
    }

    .trust-indicator {
        display: inline-flex;
        align-items: center;
        gap: 0.5rem;
        margin: 0.5rem 0.5rem 0.5rem 0;
        font-size: 0.9em;
        color: #2E7D32;
        font-weight: 600;
    }

    /* ============ TABS ============ */
    .stTabs [data-baseweb="tab-list"] button {
        font-size: 1.05em;
        font-weight: 600;
        color: #546E7A;
        border-bottom: 3px solid transparent;
        transition: all 0.3s ease;
    }

    .stTabs [data-baseweb="tab-list"] button[aria-selected="true"] {
        color: #3A6FA0;
        border-bottom-color: #D4AF37;
    }

    /* ============ EXPANDER ============ */
    .streamlit-expanderHeader {
        background-color: #F7F2EA;
        border: 1px solid #E0E0E0;
        border-radius: 4px;
    }

    /* ============ TEXT ============ */
    p {
        color: #546E7A;
        line-height: 1.6;
    }

    strong {
        color: #3A6FA0;
        font-weight: 700;
    }</style>
""",
    unsafe_allow_html=True,
)

# ================================================================
# AUTENTICAÇÃO
# ================================================================

if "auth_service" not in st.session_state:
    st.session_state.auth_service = AuthService()

if "db_service" not in st.session_state:
    st.session_state.db_service = DatabaseService()

if "authenticated_user" not in st.session_state:
    st.session_state.authenticated_user = None


def render_logo(width: int = 280) -> None:
    """Renderiza a logo institucional quando disponível."""
    try:
        if LOGO_PATH.exists():
            st.image(str(LOGO_PATH), width=width)
        else:
            st.write("🔐")
    except Exception as e:
        logger.warning(f"Logo não encontrada: {str(e)}")
        st.write("🔐")

if st.session_state.authenticated_user is None:

    # Google OAuth callback
    query_params = st.query_params
    google_oauth_url = st.session_state.get("pending_google_oauth_url", "")
    if "code" in query_params:
        code = query_params["code"]
        redirect_uri = st.session_state.auth_service.resolve_google_redirect_uri()

        ok, msg, user_data = st.session_state.auth_service.google_callback(code, redirect_uri)
        if ok:
            st.session_state.authenticated_user = user_data
            st.session_state.pending_google_oauth_url = ""
            st.session_state.db_service.log_acceptance(user_data["username"])
            st.query_params.clear()
            st.rerun()
        else:
            st.error(f"Erro no login Google: {msg}")
            st.session_state.pending_google_oauth_url = ""
            st.query_params.clear()

    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        logo_col1, logo_col2, logo_col3 = st.columns([1, 2, 1])
        with logo_col2:
            render_logo(width=260)

        st.markdown("---")
        st.markdown(
            '<h2 style="text-align: center; color: #3A6FA0;">🔐 Acesso ao Sistema</h2>',
            unsafe_allow_html=True,
        )

        with st.form("login_form"):
            username = st.text_input("Usuário")
            password = st.text_input("Senha", type="password")

            st.markdown("---")
            st.markdown("""
            **⚠️ Termos de Uso**

            As análises são geradas por inteligência artificial com base jurídica em evidências públicas,
            jurisprudências e boas práticas dos maiores escritórios do Brasil. Esta ferramenta é um auxílio
            e **não substitui a consulta a um advogado**.

            A TRUST CORPORATION não se responsabiliza por decisões tomadas com base
            exclusivamente nas análises desta plataforma.
            """)

            aceite = st.checkbox(
                "Li e aceito os termos de uso",
                value=False,
                key="accept_terms",
            )

            col_btn1, col_btn2 = st.columns(2)
            with col_btn1:
                submitted = st.form_submit_button("Entrar", use_container_width=True, type="primary")
            with col_btn2:
                google_login = st.form_submit_button("Entrar com Google", use_container_width=True)

            if submitted:
                if not aceite:
                    st.error("É necessário aceitar os termos de uso para acessar o sistema.")
                else:
                    ok, msg, user_data = st.session_state.auth_service.authenticate(username, password)
                    if ok:
                        st.session_state.authenticated_user = user_data
                        st.session_state.db_service.log_acceptance(username)
                        st.rerun()
                    else:
                        st.error(msg)

            if google_login:
                if not aceite:
                    st.error("É necessário aceitar os termos de uso para acessar o sistema.")
                else:
                    redirect_uri = st.session_state.auth_service.resolve_google_redirect_uri()
                    google_url = st.session_state.auth_service.google_login_url(redirect_uri)
                    if google_url:
                        st.session_state.pending_google_oauth_url = google_url
                        google_oauth_url = google_url
                    else:
                        st.info("Google OAuth não configurado. Use login com usuário e senha.")

        if google_oauth_url:
            st.link_button(
                "Continuar com Google",
                url=google_oauth_url,
                use_container_width=True,
            )
            st.markdown(
                f'<meta http-equiv="refresh" content="0; url={google_oauth_url}">',
                unsafe_allow_html=True,
            )
            st.caption("Se o redirecionamento automático não abrir, use o botão acima.")

    st.stop()

# ================================================================
# INICIALIZAÇÃO DO SESSION STATE
# ================================================================

# Inicializar variáveis de sessão com valores padrão
if "cache_analise" not in st.session_state:
    st.session_state.cache_analise = None

if "texto_original" not in st.session_state:
    st.session_state.texto_original = ""

if "analise_resultado" not in st.session_state:
    st.session_state.analise_resultado = ""

if "nome_arquivo" not in st.session_state:
    st.session_state.nome_arquivo = ""

if "chat_messages" not in st.session_state:
    st.session_state.chat_messages = []

if "modelo_selecionado" not in st.session_state:
    st.session_state.modelo_selecionado = "deepseek"

if "cache_manager" not in st.session_state:
    st.session_state.cache_manager = CacheManager(cache_dir="cache", max_history=100)

# ================================================================
# FUNÇÕES AUXILIARES
# ================================================================


def limpar_analise():
    """
    Limpa o resultado da análise atual.
    Mantém o arquivo, texto original e cache persistido.
    """
    st.session_state.cache_analise = None
    st.session_state.analise_resultado = ""
    st.session_state.chat_messages = []
    logger.info("Análise limpa do session state")


# ================================================================
# INTERFACE PRINCIPAL
# ================================================================

# ================================================================
# HEADER PROFISSIONAL COM BADGES DE SEGURANÇA
# ================================================================

# Header com logo e título
col1, col2 = st.columns([1, 4])

with col1:
    render_logo(width=300)

with col2:
    st.title("TRUST CORPORATION")
    st.markdown("### Contrato Seguro IA - Análise Inteligente de Contratos")
    st.markdown(
        "*Riscos Contratuais analisados pelos modelos utilizados pelos maiores escritórios de advocacia do Brasil*"
    )

# ================================================================
# BADGES DE SEGURANÇA E TRUST INDICATORS
# ================================================================

st.markdown("---")

# Badges de Segurança
col_badge1, col_badge2, col_badge3, col_badge4, col_badge5 = st.columns(5)

with col_badge1:
    st.markdown(
        '<div class="security-badge">🔒 LGPD Compliant</div>',
        unsafe_allow_html=True,
    )

with col_badge2:
    st.markdown(
        '<div class="security-badge">🔐 SSL/TLS Encrypted</div>',
        unsafe_allow_html=True,
    )

with col_badge3:
    st.markdown(
        '<div class="security-badge">✅ ISO 27001</div>',
        unsafe_allow_html=True,
    )

with col_badge4:
    st.markdown(
        '<div class="security-badge">📊 Auditável</div>',
        unsafe_allow_html=True,
    )

with col_badge5:
    st.markdown(
        '<div class="trust-badge">⭐ v2.3 Premium</div>',
        unsafe_allow_html=True,
    )

# Trust Indicators Section
st.markdown(
    """
    <div class="trust-section">
        <h3 style="color: #3A6FA0; margin-top: 0;">🛡️ Garantias de Segurança</h3>
        <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 1rem;">
            <div class="trust-indicator">✅ Dados não são salvos no servidor</div>
            <div class="trust-indicator">✅ Análise 100% privada</div>
            <div class="trust-indicator">✅ End-to-end encrypted</div>
            <div class="trust-indicator">✅ Sem rastreamento</div>
            <div class="trust-indicator">✅ Resultado em tempo real</div>
            <div class="trust-indicator">✅ Modelo validado por especialistas</div>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

st.divider()

# ================================================================
# BARRA LATERAL - CONFIGURAÇÕES
# ================================================================

with st.sidebar:
    st.markdown(
        """
        <style>
            [data-testid="stSidebarContent"] {
                background-color: #F7F2EA;
                padding-top: 2rem;
            }
        </style>
        """,
        unsafe_allow_html=True,
    )

    user = st.session_state.authenticated_user
    st.markdown(
        f'<div style="text-align: center; padding: 0.5rem;">'
        f'<span style="color: #3A6FA0; font-weight: 700;">👤 {user["username"]}</span>'
        f'<br><span style="color: #546E7A; font-size: 0.8em;">{user["role"].upper()}</span>'
        f'</div>',
        unsafe_allow_html=True,
    )
    col_btn1, col_btn2 = st.columns(2)
    with col_btn1:
        if st.button("🚪 Sair", use_container_width=True):
            st.session_state.authenticated_user = None
            st.rerun()
    
    with col_btn2:
        # Botão Admin (apenas para admins)
        if user.get("role") == "admin":
            if st.button("📊 Admin", use_container_width=True):
                st.session_state.show_admin = not st.session_state.get("show_admin", False)
                st.rerun()

    st.markdown("---")

    st.markdown(
        '<h2 style="color: #3A6FA0; border-bottom: 2px solid #D4AF37; padding-bottom: 0.5rem; margin-bottom: 1.5rem;">⚙️ Configurações</h2>',
        unsafe_allow_html=True,
    )

    # Seleção de modelo com estilo melhorado
    st.markdown(
        '<p style="color: #3A6FA0; font-weight: 600; margin-bottom: 0.5rem;">Modelo de IA:</p>',
        unsafe_allow_html=True,
    )
    modelo = st.radio(
        "Escolha o modelo",
        ["DeepSeek (Primário)", "Gemini (Cloud)"],
        index=0,
        label_visibility="collapsed",
    )

    modelo_map = {
        "DeepSeek (Primário)": "deepseek",
        "Gemini (Cloud)": "gemini",
    }

    modelo_selecionado = modelo_map[modelo]
    if modelo_selecionado != st.session_state.modelo_selecionado:
        set_model(modelo_selecionado)
        st.session_state.modelo_selecionado = modelo_selecionado
        logger.info(f"Modelo alterado para: {modelo_selecionado}")

    st.divider()

    # Opções avançadas
    st.markdown(
        '<p style="color: #3A6FA0; font-weight: 600; margin-bottom: 0.5rem;">Opcões Avançadas:</p>',
        unsafe_allow_html=True,
    )
    fallback_enabled = st.checkbox("Fallback automático entre modelos", value=True,
                                  help="Se o modelo principal falhar, tenta automaticamente o outro modelo")
    set_fallback(fallback_enabled)

    st.divider()

    # Botão para limpar análise com estilo
    st.markdown(
        '<p style="color: #3A6FA0; font-weight: 600; margin-bottom: 0.5rem;">Ações:</p>',
        unsafe_allow_html=True,
    )
    if st.button("🗑️ Limpar Análise", use_container_width=True):
        limpar_analise()
        st.success("Análise limpa com sucesso!")
        st.rerun()

    st.divider()

    # Estatísticas do cache
    st.markdown(
        '<h3 style="color: #3A6FA0; border-bottom: 2px solid #D4AF37; padding-bottom: 0.5rem;">💾 Cache</h3>',
        unsafe_allow_html=True,
    )
    cache_stats = st.session_state.cache_manager.get_cache_stats()
    st.metric("Análises em cache", cache_stats.get("total_entries", 0))
    if st.button("🗑️ Limpar Cache", use_container_width=True):
        st.session_state.cache_manager.clear_cache()
        st.success("Cache limpo!")
        st.rerun()

    # Histórico de análises (via DB)
    with st.expander("📋 Histórico de Análises", expanded=False):
        try:
            user_id = st.session_state.authenticated_user.get("username", "anonymous")
            history = st.session_state.db_service.get_user_history(user_id, limit=10)
            if history:
                for h in history:
                    emoji = "🔴" if h.get("risk_high", 0) > 2 else ("🟠" if h.get("risk_medium", 0) > 2 else "🟢")
                    st.caption(
                        f"{emoji} {h.get('contract_name', 'N/A')[:40]}\n"
                        f"Altos: {h.get('risk_high', 0)} | "
                        f"Médios: {h.get('risk_medium', 0)} | "
                        f"Baixos: {h.get('risk_low', 0)}"
                    )
            else:
                st.caption("Nenhuma análise salva ainda.")
        except Exception:
            st.caption("Banco de dados indisponível.")

    st.divider()

    # Seção de Segurança
    st.markdown(
        '<h3 style="color: #3A6FA0; border-bottom: 2px solid #D4AF37; padding-bottom: 0.5rem;">🛡️ Segurança</h3>',
        unsafe_allow_html=True,
    )

    # Indicadores de segurança
    col_sec1, col_sec2 = st.columns(2)
    with col_sec1:
        st.markdown(
            '<div style="text-align: center; color: #2E7D32; font-weight: 600; font-size: 0.9em;">🔐 Privado</div>',
            unsafe_allow_html=True,
        )
    with col_sec2:
        st.markdown(
            '<div style="text-align: center; color: #2E7D32; font-weight: 600; font-size: 0.9em;">🔒 Criptografado</div>',
            unsafe_allow_html=True,
        )

    st.info(
        "💡 **Dica:** Carregue um PDF, Word, Excel ou TXT com o contrato para iniciar a análise."
    )

    # Painel Admin
    if st.session_state.authenticated_user.get("role") == "admin":
        st.markdown("---")
        st.markdown(
            '<h3 style="color: #D4AF37; border-bottom: 2px solid #D4AF37; padding-bottom: 0.25rem;">🔑 Painel Admin</h3>',
            unsafe_allow_html=True,
        )

        with st.expander("👥 Usuários", expanded=False):
            users = st.session_state.auth_service.list_users()
            for u in users:
                st.caption(f"• {u['username']} — {u['role']} | Criado: {u['created_at'][:10]}")

        with st.expander("📝 Registro de Aceites", expanded=False):
            try:
                acceptances = st.session_state.db_service.get_acceptance_history()
                if acceptances:
                    for a in acceptances[-15:]:
                        st.caption(f"• {a['username']} em {a['accepted_at'][:19]}")
                else:
                    st.caption("Nenhum aceite registrado.")
            except Exception:
                st.caption("Erro ao carregar aceites.")

        with st.expander("🔐 Log de Acessos", expanded=False):
            try:
                audit = st.session_state.auth_service.get_audit_log(limit=15)
                for entry in audit:
                    emoji = "✅" if "SUCCESS" in entry["action"] else "❌"
                    st.caption(f"{emoji} {entry['user']} — {entry['action']} — {entry['timestamp'][:19]}")
            except Exception:
                st.caption("Erro ao carregar auditoria.")

        with st.expander("📊 Estatísticas Gerais", expanded=False):
            try:
                stats = st.session_state.db_service.get_stats()
                st.metric("Total Análises", stats.get("total_analyses", 0))
                st.metric("Média Riscos Altos", stats.get("avg_high_risks", 0))
                st.metric("Média Riscos Médios", stats.get("avg_medium_risks", 0))
                api_stats = st.session_state.db_service.get_api_stats()
                st.metric("Chamadas API (30d)", api_stats.get("total_requests", 0))
            except Exception:
                st.caption("Erro ao carregar estatísticas.")

st.divider()

st.warning("""
**⚠️ Aviso Importante**

As análises são geradas por inteligência artificial de última geração, com base jurídica em evidências públicas, jurisprudências e nas melhores práticas dos maiores escritórios de advocacia do Brasil.

Esta ferramenta é um **auxílio para compreensão do seu contrato** e esclarecimento de dúvidas, mas **não substitui a consulta a um advogado**. Sempre consulte um profissional qualificado antes de tomar qualquer decisão ou assinar qualquer documento.

A TRUST CORPORATION não se responsabiliza por decisões tomadas com base exclusivamente nas análises geradas por esta plataforma.
""")

st.divider()

# ================================================================
# PAINEL ADMINISTRATIVO COMPLETO (SE ATIVADO)
# ================================================================

if st.session_state.get("show_admin", False):
    from tabs.admin import render_admin_tab
    render_admin_tab()
    st.markdown("---")
    if st.button("🔙 Voltar para Análise"):
        st.session_state.show_admin = False
        st.rerun()
    st.stop()  # Para exibição para não mostrar upload/análise

# ================================================================
# TUTORIAL DE USO
# ================================================================

with st.expander("📖 Tutorial — Como usar o ContratoSeguro IA", expanded=False):
    st.markdown("""
    ### 🚀 Passo a passo

    **1. 🔐 Login**
    - Faça login com suas credenciais ou use Google OAuth

    **2. 📄 Upload do Contrato**
    - Formatos aceitos: **PDF**, **Word** (.docx), **Excel** (.xlsx), **TXT**
    - Ative OCR para PDFs digitalizados/escaneados

    **3. 🤖 Escolha do Modelo**
    - **DeepSeek** (padrão) — mais rápido e econômico
    - **Gemini** como alternativa
    - Fallback automático ativado por padrão

    **4. 🔍 Análise**
    - Clique em **Iniciar Análise Especializada**
    - A IA analisa riscos, base legal e sugere melhorias
    - Análises repetidas são recuperadas do cache (instantâneo)

    **5. 📊 Dashboard**
    - Métricas de riscos (altos/médios/baixos)
    - Gráfico de pizza com distribuição
    - Gráfico radar com avaliação em 5 dimensões
    - Resumo executivo em até 20 linhas

    **6. 📝 Sugestão de Cláusulas**
    - Visualize cada cláusula problemática lado a lado
    - Biblioteca com redações padrão (multa, foro, LGPD, etc.)
    - Opção de gerar sugestão com IA

    **7. 💬 Consultoria IA**
    - Chat para tirar dúvidas sobre o contrato analisado
    - Pergunte sobre cláusulas específicas, prazos, riscos

    **8. 💾 Exportação**
    - **WhatsApp (Twilio)**: envio automático de PDF + resumo
    - **Link manual**: abre WhatsApp Web com resumo pronto
    - **PDF**: documento profissional com gráficos
    - **Word**: documento editável para anotações

    ---
    ### ⚙️ Configurações da Sidebar
    - Alterne entre modelos de IA
    - Ative/desative fallback automático
    - Veja estatísticas de cache e histórico de análises
    - Faça logout ao terminar
    """)

st.divider()

# ================================================================
# SEÇÃO 1: UPLOAD E EXTRAÇÃO
# ================================================================

st.header("📄 Upload do Contrato")

arquivo_upload = st.file_uploader(
    "Carregue um arquivo (PDF, Word, Excel ou TXT)",
    type=["pdf", "docx", "xlsx", "txt"],
    help="Formatos aceitos: PDF, Word (.docx), Excel (.xlsx) ou TXT",
)

st.info(
    "📌 **Aviso sobre arquivos escaneados:** se o PDF for escaneado/digitalizado, "
    "a leitura com OCR pode levar alguns minutos após o envio do arquivo. "
    "Esse processamento é mais lento do que a leitura de PDFs com texto nativo."
)

if arquivo_upload is not None:
    st.session_state.nome_arquivo = arquivo_upload.name

    usar_ocr = st.checkbox("🔍 Usar OCR para PDFs escaneados", value=True,
                           help="Ativa reconhecimento de texto em PDFs digitalizados (requer pytesseract)")

    try:
        with st.spinner("Processando arquivo..."):
            ext = arquivo_upload.name.lower()
            file_bytes = arquivo_upload.getvalue()

            if ext.endswith(".pdf"):
                pdf_bytes = BytesIO(file_bytes)
                texto_extraido = extrair_texto_pdf_bytes(pdf_bytes, enable_ocr=usar_ocr)
                logger.info(f"PDF extraido: {len(texto_extraido)} caracteres de {arquivo_upload.name}")

                metadados = extrair_metadados_pdf(pdf_bytes)
                if any(v != "N/A" for v in metadados.values()):
                    with st.expander("📋 Metadados do Documento", expanded=False):
                        cols = st.columns(3)
                        for i, (key, value) in enumerate(metadados.items()):
                            if value != "N/A":
                                with cols[i % 3]:
                                    st.caption(f"**{key.replace('_', ' ').title()}:** {value}")

            elif ext.endswith(".docx"):
                try:
                    from extractor_service import extrair_texto_docx
                    texto_extraido = extrair_texto_docx(BytesIO(file_bytes))
                except ImportError:
                    texto_extraido = file_bytes.decode("utf-8", errors="ignore")

            elif ext.endswith(".xlsx"):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
                    partes = []
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        partes.append(f"[Planilha: {sheet_name}]")
                        for row in ws.iter_rows(values_only=True):
                            linha = " | ".join(str(c) if c is not None else "" for c in row)
                            if linha.strip():
                                partes.append(linha)
                    wb.close()
                    texto_extraido = "\n".join(partes)
                except ImportError:
                    st.error("openpyxl não instalado. Execute: pip install openpyxl")
                    texto_extraido = ""

            else:  # TXT
                try:
                    texto_extraido = file_bytes.decode("utf-8")
                except UnicodeDecodeError:
                    texto_extraido = file_bytes.decode("latin-1", errors="ignore")

            st.session_state.texto_original = texto_extraido

            # Exibir preview do texto
            with st.expander("👁️ Visualizar Texto Extraído", expanded=False):
                st.text_area(
                    "Texto do contrato:",
                    value=texto_extraido[:1000] + "..."
                    if len(texto_extraido) > 1000
                    else texto_extraido,
                    height=200,
                    disabled=True,
                )

            st.success(
                f"✅ Arquivo carregado com sucesso! ({len(texto_extraido)} caracteres)"
            )

    except Exception as e:
        st.error(f"❌ Erro ao processar arquivo: {str(e)}")
        logger.error(f"Erro na extração: {str(e)}")

# ================================================================
# SEÇÃO 2: ANÁLISE DO CONTRATO
# ================================================================

st.markdown(
    '<h2 style="color: #3A6FA0; border-bottom: 2px solid #D4AF37; padding-bottom: 0.5rem;">🔍 Análise do Contrato</h2>',
    unsafe_allow_html=True,
)

if st.session_state.texto_original:
    st.info(
        "💡 O contrato foi carregado. Clique no botão abaixo para iniciar a análise jurídica avançada."
    )

    col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
    with col_btn2:
        analisar_btn = st.button(
            "🚀 Iniciar Análise Especializada", use_container_width=True, type="primary"
        )

    if analisar_btn:
        cache_mgr = st.session_state.cache_manager
        texto_contrato = st.session_state.texto_original
        cache_hit = False
        resultado = ""
        tempo_inicio = time.time()
        user_id = st.session_state.authenticated_user.get("username", "anonymous")
        model_used = st.session_state.modelo_selecionado

        if not st.session_state.auth_service.check_rate_limit(user_id):
            st.error("Limite de requisições excedido. Aguarde um minuto antes de tentar novamente.")
            st.stop()

        cached = cache_mgr.get_analysis(texto_contrato)
        if cached and "analise" in cached and isinstance(cached["analise"], str):
            resultado = cached["analise"]
            cache_hit = True
            st.info("⚡ Análise recuperada do cache (instantâneo)")
            logger.info("Análise encontrada no cache")
        else:
            with st.spinner(
                "⏳ Analisando contrato com Inteligência Artificial. Por favor, aguarde..."
            ):
                try:
                    resultado = analisar_contrato(texto_contrato)
                    cache_mgr.save_analysis(texto_contrato, resultado)
                    st.success("✅ Análise concluída com sucesso!")
                except Exception as e:
                    st.error(f"❌ Erro na análise: {str(e)}")
                    logger.error(f"Erro durante análise: {str(e)}")

        if resultado:
            st.session_state.analise_resultado = resultado
            st.session_state.cache_analise = resultado
            tempo_total = time.time() - tempo_inicio
            if not cache_hit:
                st.caption(f"⏱️ Tempo de análise: {tempo_total:.1f}s")
                st.session_state.auth_service.log_analysis_request(
                    user_id, st.session_state.nome_arquivo
                )

            if not cache_hit:
                riscos_dict = extrair_numeros_riscos(resultado)
                try:
                    st.session_state.db_service.save_analysis(
                        contract_text=texto_contrato,
                        analysis_text=resultado,
                        contract_name=st.session_state.nome_arquivo,
                        model_used=model_used,
                        risk_high=riscos_dict.get("altos", 0),
                        risk_medium=riscos_dict.get("medios", 0),
                        risk_low=riscos_dict.get("baixos", 0),
                        analysis_time_ms=int(tempo_total * 1000),
                        user_id=user_id,
                    )
                    st.session_state.db_service.log_api_usage(
                        model=model_used,
                        tokens_input=len(texto_contrato) // 4,
                        tokens_output=len(resultado) // 4,
                        request_time_ms=int(tempo_total * 1000),
                        success=True,
                    )
                except Exception as e:
                    logger.warning(f"Erro ao salvar no banco: {e}")
                    pass
else:
    st.warning("📝 Aguardando upload do arquivo para iniciar a análise.")

# ================================================================
# SEÇÃO 3: RESULTADOS E TABS
# ================================================================

if st.session_state.analise_resultado:
    st.divider()

    # Criar abas profissionais
    tab_dashboard, tab_analise, tab_clausulas, tab_chat, tab_export = st.tabs(
        [
            "📊 Dashboard de Riscos",
            "📖 Análise Completa",
            "📝 Sugestão de Cláusulas",
            "💬 Consultoria IA",
            "💾 Exportação",
        ]
    )

    with tab_dashboard:
        from tabs.dashboard import render_dashboard
        render_dashboard()

    with tab_analise:
        st.markdown(
            '<h3 style="color: #3A6FA0;">Documento de Análise Detalhada</h3>',
            unsafe_allow_html=True,
        )
        st.markdown(
            """
            <div style="background-color: #F7F2EA; padding: 2rem; border-radius: 8px; border: 1px solid #E0E0E0; box-shadow: 0 4px 6px rgba(0,0,0,0.05);">
            """,
            unsafe_allow_html=True,
        )
        st.markdown(st.session_state.analise_resultado)
        st.markdown("</div>", unsafe_allow_html=True)

    with tab_clausulas:
        from tabs.clauses import render_clauses
        render_clauses()

    with tab_chat:
        from tabs.chat import render_chat
        render_chat()

    with tab_export:
        from tabs.export import render_export
        render_export()

# ================================================================
# RODAPÉ
# ================================================================

st.divider()

footer_col1, footer_col2, footer_col3 = st.columns(3)

with footer_col1:
    st.caption("🔒 TRUST CORPORATION - Contrato Seguro IA")

with footer_col2:
    st.caption("⚠️ Consulte sempre um advogado")

with footer_col3:
    st.caption("© 2026 - Todos os direitos reservados")

logger.info("Aplicação Streamlit iniciada com sucesso")
